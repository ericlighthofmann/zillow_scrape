import sys
from PyQt4 import QtGui, uic, QtCore
import requests
import bs4
import openpyxl
import time
import re
import os

linkList = []
maxList = []
URLList = []

#time.sleep(10)
#GUI definitions and processes
class MyWindow(QtGui.QMainWindow):
    def __init__(self):
        super(MyWindow, self).__init__()
        #.QMainWindow.__init__(self, None, QtCore.Qt.WindowStaysOnTopHint)
        uic.loadUi('mywindow.ui', self)
        self.setWindowTitle("Zillow Scrape")
        self.setWindowIcon(QtGui.QIcon('Hofdata.png'))
        self.home()

    def home(self):
        quit_button = QtGui.QPushButton("Quit", self)
        quit_button.clicked.connect(self.close_application)
        quit_button.move(500,270)

        self.progress = self.progressBar
        self.pushButton.clicked.connect(self.inform_user)
                  
        QtGui.qApp.setStyle('Plastique')
        self.show()

    def close_application(self):
        print('Its closed')
        sys.exit()

    def scrape_properties(self, zipcode, wb, sheet):
        rowNum = 1
        for link in linkList:
            rowNum += 1
            sheet.cell(row=rowNum, column=10).value = link
            res = requests.get(link)
            soup = bs4.BeautifulSoup(res.text, 'html.parser')
            # find address
            elem = soup.find('div', class_='hdp-header-description')
            elemsplit = re.split('\s\s\s\s+?', elem.text)
            address = elemsplit[0]
            sheet.cell(row=rowNum, column=1).value = address
            # find type
            elem = soup.find('div', class_='status-icon-row').text
            elem = elem.lstrip()
            sheet.cell(row=rowNum, column=2).value = elem
            #find price
            elem = soup.find('div', class_='main-row').text
            elem = elem.lstrip()
            if 'from: ' in elem:
                elem = elem.split('from: ')
                elem = elem[1]
            sheet.cell(row=rowNum, column=3).value = elem
            # find beds, baths, sqft and lot size (if applicable)
            elem = soup.find('h3')
            for e in elem:
                e = str(e)
                if 'bed' in e:
                    e = e.split('>')
                    e = e[1].split('<')
                    sheet.cell(row=rowNum, column=4).value = e[0]
                if 'bath' in e:
                    e = e.split('>')
                    e = e[1].split('<')
                    sheet.cell(row=rowNum, column=5).value = e[0]
                if 'sqft' in e:
                    e = e.split('>')
                    e = e[1].split('<')
                    sheet.cell(row=rowNum, column=6).value = e[0]
                if 'acre' in e:
                    e = e.split('>')
                    e = e[1].split('<')
                    sheet.cell(row=rowNum, column=7).value = e[0]
            # find how long its been on zillow and when its built (if applicable)
            elem = soup.find_all('ul', class_='zsg-list_square')
            for child in elem:
                li = child.find_all('li')
                for item in li:
                    if 'Built in' in item.text:
                        built = child.text.split('Built in ')
                        built = built[1]
                        built = built[0:4:]
                        sheet.cell(row=rowNum, column=9).value = built
                    if 'on Zillow' in item.text:
                        days = item.text.split(' on Zillow')
                        days = days[0]
                        daysint = ''.join(x for x in days if x.isdigit())
                        sheet.cell(row=rowNum, column=8).value = daysint
            #progress bar
            linknumber = linkList.index(link)
            linknumber = linknumber + 1
            while self.completed < (linknumber/len(linkList)*100):
                self.completed += 0.0001
                self.progress.setValue(self.completed)            
                    
        wb.save('zillow_info_ ' + zipcode + '.xlsx')
        logOutput = self.textEdit
        self.textEdit.moveCursor(QtGui.QTextCursor.End)
        logOutput.insertPlainText('All finished! Your Excel is saved in the same folder as this .exe file.')
        print ('finito')
            
        
    def collect_links(self, zipcode, wb, sheet):
        #gather links from Zillow
        URL = 'http://www.zillow.com/homes/for_sale/' + str(zipcode) + '/1_p/'
        res = requests.get(URL)
        soup = bs4.BeautifulSoup(res.text, 'html.parser')
        #finds the number of pages of results
        final_page = soup.find('ol', class_='zsg-pagination')
        for child in final_page:
            maxList.append(child.text)
        maxList[:] = (value for value in maxList if value != 'Next')
        maxList[:] = (value for value in maxList if value != '...')
        res = list(map(int, maxList))
        res2 = list(set(res))
        last_page = max(res2)
        
   
        logOutput = self.textEdit
        self.textEdit.moveCursor(QtGui.QTextCursor.End)
        logOutput.insertPlainText('There are ' + str(last_page) + ' pages of results to scrape.\n\n')

        for number in res2[0::]:
            URLList.append('http://www.zillow.com/homes/for_sale/' + str(zipcode) + '/' + str(number) + '_p/')
        for address in URLList:
            res = requests.get(address)
            soup = bs4.BeautifulSoup(res.text, 'html.parser')
            
            elem = soup.find_all('a', class_='hdp-link')
            for e in elem:
                link = 'http://www.zillow.com' + e['href']
                if link not in linkList:
                    linkList.append(link)

        logOutput.insertPlainText('We found ' + str(len(linkList)) + ' properties for ' + zipcode + '. Scraping now.' + '\n\n')
        self.scrape_properties(zipcode, wb, sheet)
            

    def inform_user(self):
        #gather zip code and log outputs
        self.pushButton.setEnabled(False)
        zipcode = self.plainTextEdit.toPlainText()
        #set up Excel spreadsheet
        filename = 'zillow_info_' + zipcode + '.xlsx'
        if not os.path.exists(filename):
            wb = openpyxl.Workbook()
            wb.save('zillow_info_ ' + zipcode + '.xlsx')
        else:
            wb = openpyxl.load_workbook(filename=filename)
        sheet = wb.active
        sheet.cell(row=1, column=1).value = 'address'
        sheet.cell(row=1, column=2).value = 'type'
        sheet.cell(row=1, column=3).value = 'price'
        sheet.cell(row=1, column=4).value = 'beds'
        sheet.cell(row=1, column=5).value = 'baths'
        sheet.cell(row=1, column=6).value = 'square feet'
        sheet.cell(row=1, column=7).value = 'lot size'
        sheet.cell(row=1, column=8).value = 'days on Zillow'
        sheet.cell(row=1, column=9).value = 'year built'
        sheet.cell(row=1, column=10).value = 'URL'
        
        self.plainTextEdit.setReadOnly(True)

        logOutput = self.textEdit
        self.textEdit.moveCursor(QtGui.QTextCursor.End)
        logOutput.insertPlainText('You entered a zip code of ' + zipcode + '\n')
        logOutput.insertPlainText('Gathering links to all visible properties in your zip code...\n')

        #progress bar
        self.completed = 0
        while self.completed < 5:
            self.completed += 0.0001
            self.progress.setValue(self.completed)

        self.collect_links(zipcode, wb, sheet)

        while self.completed < 10:
            self.completed += 0.0001
            self.progress.setValue(self.completed)

            
if __name__ == '__main__':
    app = QtGui.QApplication(sys.argv)
    window = MyWindow()
    sys.exit(app.exec_())
