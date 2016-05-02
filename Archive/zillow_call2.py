import selenium
from selenium import webdriver
import openpyxl

titleList = []
linkList = []
listing_typeList = []
priceList = []
bedsList = []
bathsList = []
sqftList = []
dozList = []

browser = webdriver.Firefox()
wb = openpyxl.load_workbook(filename = 'zillow_info.xlsx')
sheet = wb.active

def zillow_scrape(URL):
    browser.get(URL)
    elem = browser.find_elements_by_class_name('hdp-link')
    for e in elem:
        if "photos" not in e.text and "photo" not in e.text:
            title = e.text
            titleList.append(title)
            link = e.get_attribute('href')
            linkList.append(link)
        else:
            pass

    print (len(titleList))
    print (len(linkList))



URLList = ['http://www.zillow.com/homes/00802_rb/', 'http://www.zillow.com/homes/for_sale/VI-00802/58111_rid/any_days/globalrelevanceex_sort/18.429781,-64.760457,18.247611,-65.09554_rect/11_zm/2_p/',\
          'http://www.zillow.com/homes/for_sale/VI-00802/58111_rid/any_days/globalrelevanceex_sort/18.429781,-64.760457,18.247611,-65.09554_rect/11_zm/3_p/', \
          'http://www.zillow.com/homes/for_sale/VI-00802/58111_rid/any_days/globalrelevanceex_sort/18.429781,-64.760457,18.247611,-65.09554_rect/11_zm/4_p/']
for URL in URLList:
    zillow_scrape(URL)

rowNum = 2
sheet.cell(row=1, column=1).value = 'title'
sheet.cell(row=1, column=2).value = 'link'
for e in titleList:
    sheet.cell(row=rowNum, column=1).value = str(e)
    rowNum += 1
rowNum = 2
for e in linkList:
    sheet.cell(row=rowNum, column=2).value = str(e)
    rowNum += 1
rowNum = 2
wb.save('zillow_info.xlsx')
