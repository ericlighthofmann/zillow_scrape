"""Pulls data from Zillow."""

from selenium import webdriver
import openpyxl
import time

URLList = []
addressList = []
typeList = []
priceList = []
bedList = []
bathList = []
sqftList = []
lotsizeList = []
dozList = []
yearbuiltList = []
score = 0

browser = webdriver.Firefox()
wb = openpyxl.load_workbook(filename='zillow_info.xlsx')
sheet = wb.active

print('Getting info from cells...')
for row in range(2, sheet.max_row + 1):
    URL = sheet['A' + str(row)].value
    URLList.append(URL)

def zillow_scrape(URL):
    try:
        browser.get(URL)
        time.sleep(3)
        elem = browser.find_elements_by_tag_name('h1')
        for e in elem:
            addressList.append(e.text)
        elem = browser.find_element_by_class_name('status-icon-row').text
        elem = elem.strip('\" ')
        elem = elem.strip(' \"')
        typeList.append(elem)
        elemfind = browser.find_elements_by_class_name('addr_bbs')
        for elem in elemfind:
            elem = elem.get_attribute('innerHTML')
            if 'acres' in elem or 'acre' in elem:
                print(elem)
                lotsizeList.append(elem)
                bedList.append('')
                bathList.append('')
                sqftList.append('')
            elif 'bath' in elem or 'baths' in elem:
                print(elem)
                bathList.append(elem)
            elif 'bed' in elem or 'beds' in elem:
                print(elem)
                lotsizeList.append('')
                bedList.append(elem)
            elif 'sqft' in elem:
                print(elem)
                sqftList.append(elem)
            else:
                bedList.append('')
                bathList.append('')
                sqftList.append('')
        elem = browser.find_element_by_class_name('main-row').text
        priceList.append(elem)
        print(elem)
        try:
            elem = browser.find_element_by_xpath("//*[contains(text(), \
            'days on Zillow')]")
        except:
            elem = browser.find_element_by_xpath("//*[contains(text(), \
            'day on Zillow')]")
        elem = elem.text.split(' ')
        elem = elem[0]
        if elem == 'Less':
            elem = '1'
        print(elem)
        dozList.append(elem)
        elem = browser.find_element_by_xpath("//*[contains(text(), \
        'See data sources')]")
        time.sleep(1)
        elem.click()
        time.sleep(1)
        elem = browser.find_elements_by_css_selector('tr')
        for e in elem:
            if 'Year Built' in e.text:
                if e != 'Year Built: -- --':
                    e = e.text.split(': ')
                    e = e[1]
                    e = e.strip(' --')
                else:
                    e = e.text.split(': ')
                    e = e[1]
                print(e)
                yearbuiltList.append(e)
    except:
        pass

for URL in URLList:
    score += 1
    if score / len(URLList) % 10 == 0:
        print(str(score/len(URLList)) + ' % complete')
    zillow_scrape(URL)

print('Exporting to Excel...')
rowNum = 2
sheet.cell(row=1, column=2).value = 'address'
sheet.cell(row=1, column=3).value = 'type'
sheet.cell(row=1, column=4).value = 'price'
sheet.cell(row=1, column=5).value = 'beds'
sheet.cell(row=1, column=6).value = 'baths'
sheet.cell(row=1, column=7).value = 'square feet'
sheet.cell(row=1, column=8).value = 'lot size'
sheet.cell(row=1, column=9).value = 'days on Zillow'
sheet.cell(row=1, column=10).value = 'year built'
for e in addressList:
    sheet.cell(row=rowNum, column=2).value = str(e)
    rowNum += 1
rowNum = 2
for e in typeList:
    sheet.cell(row=rowNum, column=3).value = str(e)
    rowNum += 1
rowNum = 2
for e in priceList:
    sheet.cell(row=rowNum, column=4).value = str(e)
    rowNum += 1
rowNum = 2
for e in bedList:
    sheet.cell(row=rowNum, column=5).value = str(e)
    rowNum += 1
rowNum = 2
for e in bathList:
    sheet.cell(row=rowNum, column=6).value = str(e)
    rowNum += 1
rowNum = 2
for e in sqftList:
    sheet.cell(row=rowNum, column=7).value = str(e)
    rowNum += 1
rowNum = 2
for e in lotsizeList:
    sheet.cell(row=rowNum, column=8).value = str(e)
    rowNum += 1
rowNum = 2
for e in dozList:
    sheet.cell(row=rowNum, column=9).value = str(e)
    rowNum += 1
rowNum = 2
for e in yearbuiltList:
    sheet.cell(row=rowNum, column=10).value = str(e)
    rowNum += 1
rowNum = 2
wb.save('zillow_info.xlsx')
print('finito')
