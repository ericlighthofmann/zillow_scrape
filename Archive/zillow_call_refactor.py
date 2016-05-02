"""Pulls data from Zillow."""

from selenium import webdriver
import openpyxl
import time

URLList = []
addressList = []
typeList = []
priceList = []
bedList = []
bathList = []
sqftList = []
lotsizeList = []
dozList = []
yearbuiltList = []

browser = webdriver.Firefox()
wb = openpyxl.load_workbook(filename='zillow_info.xlsx')
sheet = wb.active
sheet.cell(row=1, column=2).value = 'address'
sheet.cell(row=1, column=3).value = 'type'
sheet.cell(row=1, column=4).value = 'price'
sheet.cell(row=1, column=5).value = 'beds'
sheet.cell(row=1, column=6).value = 'baths'
sheet.cell(row=1, column=7).value = 'square feet'
sheet.cell(row=1, column=8).value = 'lot size'
sheet.cell(row=1, column=9).value = 'days on Zillow'
sheet.cell(row=1, column=10).value = 'year built'

print('Getting info from cells...')
for row in range(2, sheet.max_row + 1):
    URL = sheet['A' + str(row)].value
    URLList.append(URL)

def zillow_scrape(URL, rowNum):
    try:
        browser.get(URL)
        time.sleep(3)
        # find address
        elem = browser.find_elements_by_tag_name('h1')
        for e in elem:
            sheet.cell(row=rowNum, column=2).value = e.text
        # find status
        elem = browser.find_element_by_class_name('status-icon-row').text
        elem = elem.strip('\" ')
        elem = elem.strip(' \"')
        sheet.cell(row=rowNum, column=3).value = elem
        #find price
        elem = browser.find_element_by_class_name('main-row').text
        sheet.cell(row=rowNum, column=4).value = elem
        # find bed baths sqft acres
        elemfind = browser.find_elements_by_class_name('addr_bbs')
        for elem in elemfind:
            elem = elem.get_attribute('innerHTML')
            if 'acres' in elem or 'acre' in elem:
                sheet.cell(row=rowNum, column=8).value = elem
            elif 'bath' in elem or 'baths' in elem:
                sheet.cell(row=rowNum, column=6).value = elem
            elif 'bed' in elem or 'beds' in elem:
                sheet.cell(row=rowNum, column=5).value = elem
            elif 'sqft' in elem:
                sheet.cell(row=rowNum, column=7).value = elem
            else:
                pass
        # find days on zillow
        try:
            elem = browser.find_element_by_xpath("//*[contains(text(), \
            'days on Zillow')]")
        except:
            elem = browser.find_element_by_xpath("//*[contains(text(), \
            'day on Zillow')]")
        elem = elem.text.split(' ')
        elem = elem[0]
        if elem == 'Less':
            elem = '1'
        sheet.cell(row=rowNum, column=9).value = elem
        # click see data sources
        elem = browser.find_element_by_xpath("//*[contains(text(), \
        'See data sources')]")
        time.sleep(1)
        elem.click()
        time.sleep(1)
        # find year built
        elem = browser.find_elements_by_css_selector('tr')
        for e in elem:
            if 'Year Built' in e.text:
                if e != 'Year Built: -- --':
                    e = e.text.split(': ')
                    e = e[1]
                    e = e.strip(' --')
                else:
                    e = e.text.split(': ')
                    e = e[1]
                sheet.cell(row=rowNum, column=10).value = e
        return rowNum
    except:
        return rowNum
        pass

rowNum = 2
score = 0
for URL in URLList:
    score += 1
    progress = score/len(URLList)*100
    progress = format(progress, '.2f')
    print(progress + '% complete')
    zillow_scrape(URL, rowNum)
    rowNum += 1

wb.save('zillow_info.xlsx')
print('finito')
